
import numpy as np
from openpyxl import load_workbook

#load work book
wb = load_workbook("IDCT FOUNDATION ELEVATION FOR 6-6.xlsx", data_only=True) # import your excel file here
ws = wb.active  #activates the workbook
#info extraction


Ol = 1 #length of frist level
'''
Bl = [2.6]
alpha = [0.847]
Zn = [5]'''
x = 0.2 #the step lrngth according of drawing

Vm = [] #total volume container
vi=[] #Temporary storage volume container
print('xxxxxxxxxxxxxxxxxxxxxXXXXXXXXXXxxxxxxxxxxxxxxx')

for row in range(7,35):  #based on the number of cells in calcualting range
    
    foundation_num = ws[f'B{row}'].value #gets foundation number
    steps = ws[f'L{row}'].value  #levels "Zn"
    base_length = ws[f'H{row}'].value #base length "Bl"
    final_height = ws[f'J{row}'].value # alpha 

    
        
    V = round(base_length**2 * final_height,2) #Volume of PCC base at given final height
        
    print(f'volume of base foundation - {foundation_num} : {V} m^3')

    for j in range (1,steps):
        
        Zeta =((((steps-j)-1)*2*x)+Ol)**2
        vi.append(round(Zeta,2))

    vi.insert(0,V)
    SumVol = np.array(vi)

    print(f'The volumes of PCC {foundation_num} with {steps} steps is : {SumVol}')

    TotalVol = np.sum((SumVol))
    Vm.append(TotalVol)
    del vi[:] 
print(f' Total Volumes of {foundation_num} PCC: {Vm} m^3')
print(len(Vm))
for i in range(28):
    print (f'the volume of PCC foundation {i+1} is | {round(Vm[i],2)} m^3')
